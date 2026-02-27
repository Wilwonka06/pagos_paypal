"""
Sistema de Automatización de Pagos PayPal
Integra SAP, Excel, PDFs y gestión de carpetas
"""

import os
import sys
import time
import shutil
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Tuple

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import openpyxl.utils
import fitz  # PyMuPDF

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ============================================================================
# CONFIGURACIÓN Y CONSTANTES
# ============================================================================

class Config:
    """Configuración centralizada del sistema"""
    
    # Rutas principales
    BASE_PAYPAL = Path(r"C:\Finanzas\Info Bancos\Pagos Internacionales\PAYPAL")
    RUTA_DESCARGAS = Path.home() / "Downloads"
    RUTA_MAESTRO = BASE_PAYPAL / "COURIER" / "A_Aplicación" / "Courier Internacional - Reporte Cambiario.xlsm"
    
    # Rutas de búsqueda de PDFs
    RUTAS_PDF = [
        Path(r"C:\Users\auxconmepa1\GCO\Marco Esteban Escobar Bedoya - Facturas y Guias LATAM Americanino 2025"),
        Path(r"C:\Users\auxconmepa1\GCO\Marco Esteban Escobar Bedoya - Facturas y Guias LATAM Esprit 2025"),
        Path(r"C:\Users\auxconmepa1\GCO\Marco Esteban Escobar Bedoya - Facturas y Guias LATAM Americanino 2026"),
        Path(r"C:\Users\auxconmepa1\GCO\Marco Esteban Escobar Bedoya - Facturas y Guias LATAM Esprit 2026")
    ]
    # Configuración SAP
    SAP_URL = "https://saps4h.gco.com.co/sap/bc/gui/sap/its/webgui/?sap-client=300&sap-language=ES"
    SAP_USER = "AUXCONMEPA1"
    SAP_PASSWORD = "auxilarMP01*"
    SAP_TRANSACCION = "FAGLL03"
    SAP_CUENTA = "4250060005"
    SAP_SOCIEDAD = "1000"
    
    # Configuración de Excel
    HOJA_MAESTRO = "Soporte Formulario"
        
    # Columnas del Excel (Orden exacto basado en el archivo Maestro)
    COLUMNAS_SEGUNDA_HOJA = [
        "Date", "Currency", "Gross", "Fee", "Net",
        "Prorrateo Disputa", "Prorrateo Normal", "Neto despues de prorrateo",
        "Flete", "Valor mcia",
        "Invoice Numbers", "Número guía", "Fecha del envío", 
        "Order Id Paypal", "Fecha_pago",
        "Valoración flete", "Diferencia", "Observaciones"
    ]
    
    # Timeouts
    TIMEOUT_SAP = 30
    TIMEOUT_DOWNLOAD = 60
    ACTIVAR_LOG_ARCHIVO = False

# ============================================================================
# CONFIGURACIÓN DE LOGGING
# ============================================================================

def configurar_logging():
    """Configura el sistema de logging"""
    try:
        # Forzar UTF-8 en stdout para evitar errores con emojis en Windows
        import io
        if sys.stdout.encoding != 'utf-8':
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
            
        handlers = [logging.StreamHandler(sys.stdout)]
        if getattr(Config, "ACTIVAR_LOG_ARCHIVO", False):
            log_dir = Path("logs")
            log_dir.mkdir(exist_ok=True)
            log_file = log_dir / f"automatizacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            handlers.insert(0, logging.FileHandler(log_file, encoding='utf-8'))

        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=handlers,
        )
        return logging.getLogger(__name__)
    except Exception as e:
        print(f"ERROR CRÍTICO AL CONFIGURAR LOGGING: {e}")
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
        )
        return logging.getLogger(__name__)

# ============================================================================
# FASE 1: GESTIÓN DE CARPETAS
# ============================================================================

class GestorCarpetas:
    """Gestiona la creación y organización de carpetas de pagos"""
    
    def __init__(self, base_path: Path):
        self.base_path = base_path
        self.logger = logging.getLogger(__name__)
        
        # Verificar acceso inicial
        if not self.base_path.exists():
            self.logger.error(f"RUTA BASE NO EXISTE: {self.base_path}. Por favor verifique la conexión a la red o la ruta.")
    
    def obtener_pago_pendiente_o_siguiente(self) -> int:
        """
        Busca la primera carpeta incompleta 
        """
        try:
            if not self.base_path.exists():
                raise FileNotFoundError(f"La ruta base {self.base_path} no es accesible.")

            # Buscar carpetas existentes con patrón "Pago #N"
            carpetas_existentes = []
            for d in self.base_path.iterdir():
                if d.is_dir() and d.name.startswith("Pago #"):
                    try:
                        num = int(d.name.replace("Pago #", ""))
                        carpetas_existentes.append((num, d))
                    except ValueError:
                        continue
            
            # Ordenar por número para procesar en secuencia
            carpetas_existentes.sort()

            for num, ruta in carpetas_existentes:
                # 1. Verificar si tiene Excel (Soporta .xlsx, .xlsm, .xls)
                tiene_excel = any(f.suffix.lower() in ['.xlsx', '.xlsm', '.xls'] 
                                 and not f.name.startswith("~$") 
                                 and "Soporte" not in f.name
                                 for f in ruta.iterdir() if f.is_file())
                
                # 2. Verificar si tiene soportes (PDFs)
                pdfs_en_raiz = any(f.suffix.lower() == '.pdf' for f in ruta.iterdir() if f.is_file())
                
                ruta_soporte = ruta / "Soporte"
                pdfs_en_soporte = False
                if ruta_soporte.exists():
                    pdfs_en_soporte = any(f.suffix.lower() == '.pdf' for f in ruta_soporte.iterdir() if f.is_file())

                tiene_soportes = pdfs_en_raiz or pdfs_en_soporte

                if not tiene_excel or not tiene_soportes:
                    razon = []
                    if not tiene_excel: razon.append("falta Excel")
                    if not tiene_soportes: razon.append("faltan soportes")
                    self.logger.info(f"CARPETA INCOMPLETA DETECTADA: Pago #{num} ({', '.join(razon)}). Usando este número.")
                    return num

            # Si todas están completas, obtener el siguiente número
            siguiente = (carpetas_existentes[-1][0] + 1) if carpetas_existentes else 1
            self.logger.info(f"Todas las carpetas están completas. Siguiente número: {siguiente}")
            return siguiente

        except Exception as e:
            self.logger.error(f"ERROR AL BUSCAR PAGO PENDIENTE: {str(e)}")
            return 1

    def obtener_siguiente_numero_pago(self) -> int:
        # Mantener por compatibilidad si es necesario, pero redirigir a la nueva lógica
        return self.obtener_pago_pendiente_o_siguiente()
    
    def crear_estructura_pago(self, numero_pago: int) -> Tuple[Path, Path]:
        """
        Crea la estructura de carpetas para un pago
        """
        try:
            carpeta_pago = self.base_path / f"Pago #{numero_pago}"
            carpeta_soporte = carpeta_pago / "Soporte"
            
            self.logger.info(f"Creando carpetas en: {carpeta_pago}")
            carpeta_pago.mkdir(parents=True, exist_ok=True)
            carpeta_soporte.mkdir(exist_ok=True)
            
            self.logger.info(f"Estructura creada exitosamente: {carpeta_pago}")
            return carpeta_pago, carpeta_soporte
        
        except PermissionError:
            self.logger.error(f"ERROR DE PERMISOS: No se pudieron crear las carpetas en {self.base_path}. Verifique sus permisos de escritura.")
            raise
        except Exception as e:
            self.logger.error(f"ERROR AL CREAR ESTRUCTURA DE CARPETAS: {str(e)}")
            raise

# ============================================================================
# FASE 2: DESCARGA DESDE SAP
# ============================================================================

class DescargadorSAP:
    """Maneja la descarga de archivos desde SAP"""
    
    def __init__(self):
        self.driver = None
        self.logger = logging.getLogger(__name__)
        self.download_path = Config.RUTA_DESCARGAS
    
    def configurar_chrome(self) -> webdriver.Chrome:
        """Configura el navegador Google Chrome con opciones personalizadas"""
        try:
            from selenium.webdriver.chrome.service import Service
            from subprocess import CREATE_NO_WINDOW
            
            options = webdriver.ChromeOptions()
            
            # Desactivar logs innecesarios de la consola
            options.add_argument("--log-level=3")
            options.add_experimental_option('excludeSwitches', ['enable-logging'])

            prefs = {
                "download.default_directory": str(self.download_path),
                "download.prompt_for_download": False,
                "safebrowsing.enabled": True,
            }
            options.add_experimental_option("prefs", prefs)
            
            # Configurar el servicio para evitar ventanas de consola y ruidos
            service = Service()
            service.creation_flags = CREATE_NO_WINDOW
            
            driver = webdriver.Chrome(options=options, service=service)
            self.logger.info("Navegador Google Chrome configurado correctamente")
            return driver
        
        except Exception as e:
            self.logger.error(f"Error al configurar Google Chrome: {e}")
            raise
    
    def esperar_descarga(self, timeout: int = Config.TIMEOUT_DOWNLOAD, patron_alternativo: str = None) -> Optional[Path]:
        """Espera a que se complete la descarga y retorna el archivo"""
        try:
            inicio = time.time()
            self.logger.info(f"Iniciando espera de descarga (timeout: {timeout}s)...")
            
            while time.time() - inicio < timeout:
                # Buscar archivos .xlsx recién descargados
                archivos = list(self.download_path.glob("EXPORT_*.xlsx"))
                
                # Si se proporcionó un patrón alternativo (ej: "pago 12*")
                if patron_alternativo:
                    archivos.extend(list(self.download_path.glob(patron_alternativo)))
                
                # Verificar que no haya archivos temporales
                archivos_temp = list(self.download_path.glob("*.part"))
                archivos_temp.extend(list(self.download_path.glob("*.crdownload")))
                
                if archivos and not archivos_temp:
                    # Ordenar por fecha de modificación
                    archivo_mas_reciente = max(archivos, key=lambda x: x.stat().st_mtime)
                    
                    # Verificar que el archivo tenga al menos 1KB
                    # Y que haya sido creado/modificado en los últimos 5 minutos
                    antiguedad_segundos = time.time() - archivo_mas_reciente.stat().st_mtime
                    if archivo_mas_reciente.stat().st_size > 1024 and antiguedad_segundos < 300:
                        self.logger.info(f"Archivo detectado y verificado: {archivo_mas_reciente}")
                        return archivo_mas_reciente
                    elif archivo_mas_reciente.stat().st_size > 1024 and antiguedad_segundos >= 300:
                        self.logger.warning(f"Archivo encontrado pero es antiguo ({int(antiguedad_segundos)}s). Ignorando: {archivo_mas_reciente.name}")
                
                time.sleep(2)
            
            patron_error = patron_alternativo if patron_alternativo else "EXPORT_*.xlsx"
            self.logger.error(f"TIEMPO EXCEDIDO: No se encontró el archivo {patron_error} en {self.download_path} después de {timeout} segundos.")
            return None
        
        except Exception as e:
            self.logger.error(f"ERROR DURANTE LA ESPERA DE DESCARGA: {str(e)}")
            return None
    
    def _esperar_y_hacer(self, wait, by, selector, accion="presence", valor=None, descripcion=""):
        """Método auxiliar para esperar elementos con mensajes de error claros"""
        try:
            if accion == "presence":
                elemento = wait.until(EC.presence_of_element_located((by, selector)))
            elif accion == "clickable":
                elemento = wait.until(EC.element_to_be_clickable((by, selector)))
            else:
                raise ValueError(f"Acción no reconocida: {accion}")
            
            if valor is not None:
                elemento.clear() if hasattr(elemento, 'clear') else None
                elemento.send_keys(valor)
            
            return elemento
        except TimeoutException:
            msg = f"ERROR DE TIEMPO: No se encontró o no está disponible: '{descripcion}' (Selector: {selector})"
            self.logger.error(msg)
            raise Exception(msg)
        except Exception as e:
            msg = f"ERROR EN PASO '{descripcion}': {str(e)}"
            self.logger.error(msg)
            raise Exception(msg)

    def descargar_reporte_sap(self, numero_pago: int) -> Optional[Path]:
        """
        Ejecuta el proceso completo de descarga desde SAP
        """
        try:
            self.numero_pago_actual = numero_pago
            self.driver = self.configurar_chrome()
            wait = WebDriverWait(self.driver, Config.TIMEOUT_SAP)
            
            # 1. Acceder a SAP
            self.logger.info("Accediendo a la URL de SAP...")
            try:
                self.driver.get(Config.SAP_URL)
                self.driver.set_window_size(1137, 694)
            except Exception as e:
                self.logger.error(f"Fallo al cargar la URL de SAP: {e}")
                return None
            
            # 2. Login
            self.logger.info("Esperando pantalla de login...")
            self._esperar_y_hacer(wait, By.ID, "sap-user", "presence", Config.SAP_USER, "Campo Usuario")
            self._esperar_y_hacer(wait, By.ID, "sap-password", "presence", Config.SAP_PASSWORD, "Campo Contraseña")
            self.driver.find_element(By.ID, "sap-password").send_keys(Keys.ENTER)
            
            time.sleep(3)
            
            # 3. Navegar a transacción
            self.logger.info(f"Navegando a transacción {Config.SAP_TRANSACCION}...")
            toolbar = self._esperar_y_hacer(wait, By.ID, "ToolbarOkCode", "presence", descripcion="Barra de transacciones (ToolbarOkCode)")
            toolbar.send_keys(Config.SAP_TRANSACCION)
            toolbar.send_keys(Keys.ENTER)
            
            time.sleep(2)
            
            # 4. Completar campos
            self.logger.info(f"Completando formulario de {Config.SAP_TRANSACCION}...")
            
            # Campo 1: Cuenta
            self._esperar_y_hacer(wait, By.ID, "M0:46:::1:34", "presence", Config.SAP_CUENTA, "Campo Cuenta de Mayor")
            
            # Campo 2: Sociedad
            self._esperar_y_hacer(wait, By.ID, "M0:46:::2:34", "presence", Config.SAP_SOCIEDAD, "Campo Sociedad")
            
            # 5. Ejecutar
            self.logger.info("Ejecutando la consulta (botón F8)...")
            self.driver.execute_script("window.scrollTo(0,0)")
            ejecutar_btn = self._esperar_y_hacer(wait, By.ID, "M0:50::btn[8]", "clickable", descripcion="Botón Ejecutar (F8)")
            ejecutar_btn.click()
            
            # Nombre del archivo dinámico para usar en las instrucciones
            nombre_archivo = f"pago {getattr(self, 'numero_pago_actual', 1)}"

            # ESPERA 1: Carga de la tabla inicial
            self.logger.info("Esperando carga de la tabla de resultados (15s)...")
            time.sleep(15)

            # Paso 1: Enviar Mayus + F4 para abrir la ventana de exportación
            self.logger.info("Enviando Mayus + F4 para abrir la ventana de exportación...")
            try:
                cuerpo = self.driver.find_element(By.TAG_NAME, "body")
                cuerpo.click()
                time.sleep(2)
                cuerpo.send_keys(Keys.SHIFT + Keys.F4)
                self.logger.info("Atajo Mayus + F4 enviado.")
            except Exception as e:
                self.logger.warning(f"No se pudo enviar Mayus+F4 automáticamente: {e}")

            # Paso 2: Rellenar el nombre del archivo y generar
            self.logger.info("Intentando rellenar el nombre del archivo y generar reporte...")
            try:
                campo_nombre = WebDriverWait(self.driver, 20).until(
                    EC.element_to_be_clickable((By.ID, "M1:46:1::1:17"))
                )
                campo_nombre.clear()
                campo_nombre.send_keys(nombre_archivo)

                generar_btn = self.driver.find_element(By.ID, "M1:48::btn[20]")
                generar_btn.click()
                self.logger.info(f"Nombre '{nombre_archivo}' enviado y botón Generar presionado.")
            except Exception as e:
                self.logger.warning(f"No se pudo completar el llenado del nombre/generación automática: {e}")
            
            # Confirmar descarga si aparece el diálogo de SAP
            try:
                confirmar_btn = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.ID, "UpDownDialogChoose"))
                )
                confirmar_btn.click()
            except:
                pass
            
            # 7. Esperar descarga
            archivo = self.esperar_descarga(patron_alternativo=f"{nombre_archivo}*")
            return archivo
        
        except Exception as e:
            self.logger.error(f"FALLO CRÍTICO EN DESCARGA SAP: {str(e)}")
            return None
        
        finally:
            if self.driver:
                try:
                    self.driver.quit()
                except:
                    pass

# ============================================================================
# FASE 3: PROCESAMIENTO DE EXCEL
# ============================================================================

class ProcesadorExcel:
    """Procesa archivos Excel y realiza transformaciones"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def buscar_archivo_pago_en_descargas(self, numero_pago: int) -> Optional[Path]:
        """
        Busca un archivo que contenga 'pago [numero]' en la carpeta de descargas
        """
        try:
            descargas = Config.RUTA_DESCARGAS
            self.logger.info(f"Buscando archivo para pago {numero_pago} en {descargas}...")
            
            # Patrones posibles: "pago 12", "pago#12", "pago #12", "Pago 12", "Pago #12"
            patrones = [
                f"pago {numero_pago}*",
                f"pago#{numero_pago}*",
                f"pago #{numero_pago}*",
                f"Pago {numero_pago}*",
                f"Pago#{numero_pago}*",
                f"Pago #{numero_pago}*",
                f"*pago*{numero_pago}*"
            ]
            
            for patron in patrones:
                archivos = list(descargas.glob(patron))
                # Filtrar solo archivos (no carpetas) y preferir .xlsx o .xls
                archivos = [f for f in archivos if f.is_file() and f.suffix.lower() in ['.xlsx', '.xls']]
                
                if archivos:
                    # Retornar el más reciente si hay varios
                    archivos.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                    self.logger.info(f"¡Archivo encontrado!: {archivos[0].name}")
                    return archivos[0]
            
            self.logger.warning(f"No se encontró ningún archivo para el pago {numero_pago} en Descargas.")
            return None
            
        except Exception as e:
            self.logger.error(f"Error al buscar archivo en descargas: {e}")
            return None

    def mover_y_renombrar_descarga(self, archivo_descarga: Path, 
                                   carpeta_destino: Path, 
                                   numero_pago: int) -> Path:
        """Mueve el archivo descargado a la carpeta del pago y lo renombra"""
        try:
            if not archivo_descarga.exists():
                raise FileNotFoundError(f"El archivo de descarga no existe: {archivo_descarga}")

            # Nuevo nombre con número de pago
            fecha = datetime.now().strftime("%Y%m%d")
            nuevo_nombre = f"EXPORT_{fecha}_Pago#{numero_pago}.xlsx"
            ruta_destino = carpeta_destino / nuevo_nombre
            
            # Mover archivo
            self.logger.info(f"Moviendo archivo de {archivo_descarga.name} a {ruta_destino}")
            shutil.move(str(archivo_descarga), str(ruta_destino))
            
            return ruta_destino
        
        except Exception as e:
            self.logger.error(f"ERROR AL MOVER/RENOMBRAR ARCHIVO EXCEL: {str(e)}")
            raise
    
    def reorganizar_columnas_primera_hoja(self, archivo: Path):
        """Mueve la columna 'Referencia' a la primera posición"""
        try:
            self.logger.info(f"Reorganizando columnas en {archivo.name}...")
            # Cargar con openpyxl para mantener formato
            wb = load_workbook(archivo)
            ws = wb.active
            
            # Encontrar columna "Referencia"
            headers = [cell.value for cell in ws[1]]
            
            if "Referencia" not in headers:
                self.logger.error(f"ERROR: No se encontró la columna 'Referencia' en el archivo {archivo.name}. Las columnas encontradas son: {headers}")
                return
            
            idx_referencia = headers.index("Referencia") + 1  # openpyxl usa índices desde 1
            
            # Si ya está en la primera posición, no hacer nada
            if idx_referencia == 1:
                self.logger.info("La columna 'Referencia' ya está en la primera posición.")
                return
            
            # Insertar nueva columna al inicio
            ws.insert_cols(1)
            
            # Copiar datos de la columna Referencia
            for row in range(1, ws.max_row + 1):
                valor = ws.cell(row, idx_referencia + 1).value  # +1 por la inserción
                ws.cell(row, 1).value = valor
            
            # Eliminar la columna antigua
            ws.delete_cols(idx_referencia + 1)
            
            wb.save(archivo)
            self.logger.info("Columna 'Referencia' movida a la primera posición correctamente.")
        
        except Exception as e:
            self.logger.error(f"ERROR AL REORGANIZAR COLUMNAS EN EXCEL: {str(e)}")
            raise
    
    def crear_segunda_hoja(self, archivo_principal: Path, 
                          archivo_maestro: Path) -> pd.DataFrame:
        """
        Crea la segunda hoja filtrando datos del maestro por mes actual
        """
        try:
            # Leer archivo maestro
            self.logger.info(f"Leyendo archivo maestro: {archivo_maestro}")
            if not archivo_maestro.exists():
                raise FileNotFoundError(f"No se encontró el archivo maestro en: {archivo_maestro}")

            # 1. Identificar la hoja correcta
            nombre_hoja = Config.HOJA_MAESTRO
            try:
                # Leer con dtype=object para preservar tipos originales de Excel y evitar conversiones automáticas
                df_maestro = pd.read_excel(archivo_maestro, sheet_name=nombre_hoja, engine='openpyxl', dtype=object)
                
                # Si las primeras columnas son 'Unnamed', es muy probable que el encabezado esté más abajo
                if any('Unnamed' in str(col) for col in df_maestro.columns[:3]):
                    self.logger.info("Detectados encabezados 'Unnamed'. Reintentando lectura desde la fila 2...")
                    df_maestro = pd.read_excel(archivo_maestro, sheet_name=nombre_hoja, engine='openpyxl', header=1, dtype=object)
                
                # Si la primera columna sigue siendo Unnamed (columna A vacía), la eliminamos
                if 'Unnamed: 0' in df_maestro.columns:
                    self.logger.info("Eliminando primera columna vacía (Columna A)...")
                    df_maestro = df_maestro.drop(columns=['Unnamed: 0'])

                self.logger.info(f"Hoja '{nombre_hoja}' leída correctamente.")
            except Exception as e:
                self.logger.error(f"ERROR: No se pudo leer la hoja '{nombre_hoja}' en el archivo maestro: {e}")
                self.logger.info("Intentando listar hojas disponibles...")
                xl = pd.ExcelFile(archivo_maestro)
                self.logger.info(f"Hojas encontradas: {xl.sheet_names}")
                raise

            # 2. Normalizar nombres de columnas (quitar espacios y poner en minúsculas para búsqueda flexible)
            df_maestro.columns = [str(col).strip() for col in df_maestro.columns]
            columnas_normalizadas = {col.lower(): col for col in df_maestro.columns}
            
            # Obtener mes y año actual
            mes_actual = datetime.now().month
            año_actual = datetime.now().year
            
            # 3. Filtrar por Fecha_pago del mes actual
            COL_FECHA_PAGO = 'Fecha_pago'

            if COL_FECHA_PAGO not in df_maestro.columns:
                # La columna no existe en absoluto — error estructural del maestro
                self.logger.error(
                    f"ERROR CRÍTICO: La columna '{COL_FECHA_PAGO}' no existe en la hoja "
                    f"'{nombre_hoja}' del archivo maestro.\n"
                    f"Columnas disponibles: {list(df_maestro.columns)}"
                )
                raise ValueError(
                    f"No se encontró la columna '{COL_FECHA_PAGO}' en el maestro. "
                    f"Verifique que el archivo sea el correcto."
                )

            self.logger.info(f"Columna de fecha identificada como: '{COL_FECHA_PAGO}'")
            df_maestro[COL_FECHA_PAGO] = pd.to_datetime(df_maestro[COL_FECHA_PAGO], errors='coerce')

            # Contar cuántos valores no-nulos tiene Fecha_pago en total
            total_con_fecha = df_maestro[COL_FECHA_PAGO].notna().sum()
            self.logger.info(f"Registros con '{COL_FECHA_PAGO}' llenada en el maestro: {total_con_fecha} / {len(df_maestro)}")

            if total_con_fecha == 0:
                # La columna existe pero está completamente vacía
                self.logger.error(
                    f"\n{'!'*60}\n"
                    f"PROCESO DETENIDO: La columna '{COL_FECHA_PAGO}' existe pero no tiene\n"
                    f"ningún valor llenado en el archivo maestro.\n\n"
                    f"ACCIÓN REQUERIDA:\n"
                    f"  1. Abra el archivo maestro:\n"
                    f"     {archivo_maestro}\n"
                    f"  2. En la hoja '{nombre_hoja}', columna '{COL_FECHA_PAGO}',\n"
                    f"     ingrese la fecha del pago ({mes_actual:02d}/{año_actual}) para cada\n"
                    f"     registro que corresponda a este mes.\n"
                    f"  3. Guarde el archivo y vuelva a ejecutar el script.\n"
                    f"{'!'*60}"
                )
                raise ValueError(
                    f"'{COL_FECHA_PAGO}' está vacía. Llene las fechas del mes "
                    f"{mes_actual:02d}/{año_actual} en el maestro y vuelva a ejecutar."
                )

            df_filtrado = df_maestro[
                (df_maestro[COL_FECHA_PAGO].dt.month == mes_actual) &
                (df_maestro[COL_FECHA_PAGO].dt.year == año_actual)
            ].copy()

            self.logger.info(f"Registros encontrados para el mes {mes_actual:02d}/{año_actual}: {len(df_filtrado)}")

            if len(df_filtrado) == 0:
                # La columna tiene datos pero ninguno coincide con el mes actual
                meses_disponibles = (
                    df_maestro[COL_FECHA_PAGO]
                    .dropna()
                    .dt.to_period('M')
                    .value_counts()
                    .sort_index()
                    .to_string()
                )
                self.logger.error(
                    f"\n{'!'*60}\n"
                    f"PROCESO DETENIDO: No hay registros con '{COL_FECHA_PAGO}' en\n"
                    f"{mes_actual:02d}/{año_actual} (mes actual del dispositivo).\n\n"
                    f"Meses con registros disponibles en el maestro:\n{meses_disponibles}\n\n"
                    f"ACCIÓN REQUERIDA:\n"
                    f"  1. Abra el archivo maestro:\n"
                    f"     {archivo_maestro}\n"
                    f"  2. En la hoja '{nombre_hoja}', columna '{COL_FECHA_PAGO}',\n"
                    f"     ingrese la fecha del pago ({mes_actual:02d}/{año_actual}) para los\n"
                    f"     registros de este mes y vuelva a ejecutar.\n"
                    f"{'!'*60}"
                )
                raise ValueError(
                    f"No hay registros con '{COL_FECHA_PAGO}' en {mes_actual:02d}/{año_actual}. "
                    f"Llene las fechas del mes actual en el maestro y vuelva a ejecutar."
                )
            
            df_final = pd.DataFrame(columns=Config.COLUMNAS_SEGUNDA_HOJA)
            
            for col_requerida in Config.COLUMNAS_SEGUNDA_HOJA:
                col_requerida_lower = col_requerida.lower()
                
                # Búsqueda flexible para columnas específicas
                if col_requerida_lower == "neto despues de prorrateo":
                    # Intentar con y sin coma
                    for posible in ["neto despues de prorrateo", "neto, despues de prorrateo"]:
                        if posible in columnas_normalizadas:
                            col_maestro = columnas_normalizadas[posible]
                            serie_datos = df_filtrado[col_maestro]
                            # Asegurar que sea Serie incluso si hay duplicados
                            if isinstance(serie_datos, pd.DataFrame):
                                serie_datos = serie_datos.iloc[:, 0]
                            df_final[col_requerida] = serie_datos.values
                            break
                elif col_requerida_lower in columnas_normalizadas:
                    col_maestro = columnas_normalizadas[col_requerida_lower]
                    serie_datos = df_filtrado[col_maestro]
                    if isinstance(serie_datos, pd.DataFrame):
                        serie_datos = serie_datos.iloc[:, 0]
                    df_final[col_requerida] = serie_datos.values
                else:
                    df_final[col_requerida] = None

            col_invoice = "Invoice Numbers"
            columnas_a_llenar = ["Order Id Paypal", "Número guía", "Gross", "Fee", "Flete", "Valor mcia"]

            if col_invoice in df_final.columns:
                for invoice_val, grupo in df_final.groupby(col_invoice):
                    if pd.isna(invoice_val) or str(invoice_val).strip() == "":
                        continue
                    for col in columnas_a_llenar:
                        if col not in df_final.columns:
                            continue
                        serie_fuente = grupo[col]
                        serie_fuente = serie_fuente[serie_fuente.notna()]
                        serie_fuente = serie_fuente[serie_fuente.astype(str).str.strip() != ""]
                        if serie_fuente.empty:
                            continue
                        valor = serie_fuente.iloc[0]
                        mask_faltante = (df_final[col_invoice] == invoice_val) & (
                            df_final[col].isna() | (df_final[col].astype(str).str.strip() == "")
                        )
                        df_final.loc[mask_faltante, col] = valor

            # Evitar duplicar "Valor mcia" cuando hay pagos divididos (mismo Invoice Numbers)
            if "Valor mcia" in df_final.columns and col_invoice in df_final.columns:
                df_final["Valor mcia"] = pd.to_numeric(df_final["Valor mcia"], errors="coerce")
                for invoice_val, grupo in df_final.groupby(col_invoice):
                    if pd.isna(invoice_val) or str(invoice_val).strip() == "":
                        continue
                    idxs = grupo.index.tolist()
                    if len(idxs) <= 1:
                        continue
                    vals = df_final.loc[idxs, "Valor mcia"].fillna(0)
                    nz = vals[vals != 0]
                    if nz.empty:
                        continue
                    keep_idx = nz.index[0]
                    to_zero = [i for i in idxs if i != keep_idx]
                    df_final.loc[to_zero, "Valor mcia"] = 0

            # 5. Formatear columnas de fecha y IDs, y detectar "Próximo pago"
            columnas_fecha = ["Date", "Fecha del envío", "Fecha_pago"]
            for col in columnas_fecha:
                if col in df_final.columns:
                    # Convertir a datetime y luego a string con formato DD/MM/YYYY para evitar horas en Excel
                    df_final[col] = pd.to_datetime(df_final[col], errors='coerce').dt.strftime('%d/%m/%Y')
                    # Reemplazar 'NaT' (resultado de errores en to_datetime) por None o vacío
                    df_final[col] = df_final[col].replace('NaT', None)
            
            # Detectar registros de "Próximo pago"
            # Un registro es parcial si Net > 0 pero Gross y Fee están vacíos/nulos
            if "Observaciones" in df_final.columns:
                # Obtener nombres reales de columnas mapeadas para evitar problemas de mayúsculas
                col_net = columnas_normalizadas.get('net')
                col_gross = columnas_normalizadas.get('gross')
                col_fee = columnas_normalizadas.get('fee')

                def obtener_serie_segura(df, col_name):
                    if col_name and col_name in df.columns:
                        s = df[col_name]
                        if isinstance(s, pd.DataFrame):
                            s = s.iloc[:, 0]
                        return pd.to_numeric(s, errors='coerce').fillna(0)
                    return pd.Series(0, index=df.index)

                net_vals = obtener_serie_segura(df_filtrado, col_net)
                gross_vals = obtener_serie_segura(df_filtrado, col_gross)
                fee_vals = obtener_serie_segura(df_filtrado, col_fee)
                
                # Máscara: Net > 0 Y Gross == 0 Y Fee == 0
                es_parcial = (net_vals > 0) & (gross_vals == 0) & (fee_vals == 0)
                
                # Resetear índice para asegurar alineación con df_final
                es_parcial_reset = es_parcial.reset_index(drop=True)
                df_final.loc[es_parcial_reset, "Observaciones"] = "Proximo pago"
                
                conteo_parciales = es_parcial_reset.sum()
                if conteo_parciales > 0:
                    self.logger.info(f"Se detectaron {conteo_parciales} registros de 'Próximo pago' (parciales).")
            
            # Asegurar que IDs se manejen como string para evitar notación científica y pérdida de precisión
            columnas_id = ["Order Id Paypal", "Invoice Numbers", "Número guía"]
            
            def clean_id(x):
                if pd.isna(x) or str(x).strip().lower() in ["", "nan"]:
                    return ""
                # Si es float, intentar convertir a int para quitar .0
                if isinstance(x, float):
                    if x.is_integer():
                        return str(int(x))
                    return "{:.0f}".format(x) # Evitar notación científica
                return str(x).strip()

            for col_id in columnas_id:
                if col_id in df_final.columns:
                    df_final[col_id] = df_final[col_id].apply(clean_id)

            return df_final
        
        except Exception as e:
            self.logger.error(f"ERROR AL PROCESAR SEGUNDA HOJA DESDE MAESTRO: {str(e)}")
            raise
    
    def calcular_mon_grupo_y_diferencia(self, archivo_principal: Path, df_segunda_hoja: pd.DataFrame) -> pd.DataFrame:
        """
        Calcula Comparación flete (desde SAP) y Resultado comparación usando la primera hoja
        """
        try:
            self.logger.info("Calculando valores de Comparación flete y Resultado comparación...")
            # Leer primera hoja (Data SAP) con dtype=object para IDs
            df_sap = pd.read_excel(archivo_principal, sheet_name=0, engine='openpyxl', dtype=object)
            
            # Limpiar nombres de columnas SAP
            df_sap.columns = [str(col).strip() for col in df_sap.columns]
            
            # Verificar que existan las columnas necesarias en SAP
            col_ref_sap = 'Referencia'
            col_valor_sap = 'Mon.grupo/Valoración grupo'
            
            if col_ref_sap not in df_sap.columns:
                self.logger.error(f"ERROR: No se encontró columna '{col_ref_sap}' en la hoja de SAP.")
                return df_segunda_hoja
            
            if col_valor_sap not in df_sap.columns:
                self.logger.error(f"ERROR: No se encontró columna '{col_valor_sap}' en la hoja de SAP.")
                return df_segunda_hoja
            
            # Crear diccionario de Referencia (SAP) -> Valor
            df_sap[col_ref_sap] = df_sap[col_ref_sap].astype(str).str.strip()
            dict_valores_sap = df_sap.set_index(col_ref_sap)[col_valor_sap].to_dict()
            
            # Mapear a la segunda hoja usando Invoice Numbers
            if 'Invoice Numbers' in df_segunda_hoja.columns:
                # Limpiar invoice numbers para el mapeo
                df_segunda_hoja['Invoice_Clean'] = df_segunda_hoja['Invoice Numbers'].astype(str).str.strip()
                
                # Comparación flete = Valor de SAP (Mon.grupo)
                df_segunda_hoja['Valoración flete'] = df_segunda_hoja['Invoice_Clean'].map(dict_valores_sap)
                
                # Resultado comparación = Flete + Comparación flete
                if 'Flete' in df_segunda_hoja.columns:
                    flete = pd.to_numeric(df_segunda_hoja['Flete'], errors='coerce').fillna(0)
                    comp_flete = pd.to_numeric(df_segunda_hoja['Valoración flete'], errors='coerce').fillna(0)
                    df_segunda_hoja['Diferencia'] = flete + comp_flete
                
                df_segunda_hoja.drop(columns=['Invoice_Clean'], inplace=True)
                
                conteo_mapeados = df_segunda_hoja['Valoración flete'].notna().sum()
                self.logger.info(f"Cálculos de valoración completados. {conteo_mapeados} registros vinculados.")
            
            return df_segunda_hoja
        
        except Exception as e:
            self.logger.error(f"ERROR EN CÁLCULOS DE MONEDA/DIFERENCIA: {str(e)}")
            return df_segunda_hoja
    
    def guardar_excel_con_dos_hojas(self, archivo: Path, df_segunda_hoja: pd.DataFrame):
        """Guarda el Excel con ambas hojas, aplica fórmulas dinámicas, totales y estilos"""
        try:
            self.logger.info(f"Guardando cambios finales en {archivo.name} con fórmulas dinámicas y totales...")
            
            # Identificar registros de "Proximo pago" (aquellos que NO tienen factura/guia y su Net > 0)
            # O basándonos en la lógica del maestro: registros parciales.
            df_normales = df_segunda_hoja[df_segunda_hoja['Observaciones'] != 'Proximo pago'].copy()
            df_proximos = df_segunda_hoja[df_segunda_hoja['Observaciones'] == 'Proximo pago'].copy()

            # 1. Guardar el DataFrame con pandas (solo los datos base)
            with pd.ExcelWriter(archivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df_normales.to_excel(writer, sheet_name='Reporte Procesado', index=False)
            
            # 2. Reabrir con openpyxl para insertar fórmulas, totales y registros parciales
            wb = load_workbook(archivo)
            ws = wb['Reporte Procesado']
            
            # Identificar índices de columnas
            headers = [cell.value for cell in ws[1]]
            cols_totales = [
                "Gross", "Fee", "Net",
                "Prorrateo Disputa", "Prorrateo Normal", "Neto despues de prorrateo",
                "Flete", "Valor mcia"
            ]
            indices_totales = {col: headers.index(col) + 1 for col in cols_totales if col in headers}
            
            idx_flete = headers.index('Flete') + 1 if 'Flete' in headers else None
            idx_valoracion = headers.index('Valoración flete') + 1 if 'Valoración flete' in headers else None
            idx_diferencia = headers.index('Diferencia') + 1 if 'Diferencia' in headers else None
            idx_observaciones = headers.index('Observaciones') + 1 if 'Observaciones' in headers else None

            last_row_data = ws.max_row

            # 3. Aplicar fórmulas dinámicas fila por fila (Diferencia)
            if idx_flete and idx_valoracion and idx_diferencia:
                col_flete_letra = get_column_letter(idx_flete)
                col_valoracion_letra = get_column_letter(idx_valoracion)
                col_diferencia_letra = get_column_letter(idx_diferencia)
                
                for row_idx in range(2, last_row_data + 1):
                    formula_diff = f"={col_flete_letra}{row_idx}+{col_valoracion_letra}{row_idx}"
                    ws.cell(row=row_idx, column=idx_diferencia).value = formula_diff

            # 4. Insertar fila de TOTALES
            fila_totales = last_row_data + 1
            for col_name, col_idx in indices_totales.items():
                letra = get_column_letter(col_idx)
                # Fórmula SUM desde fila 2 hasta la última de datos
                formula_sum = f"=SUM({letra}2:{letra}{last_row_data})"
                ws.cell(row=fila_totales, column=col_idx).value = formula_sum
                ws.cell(row=fila_totales, column=col_idx).font = Font(bold=True)

            # 5. Insertar registros de "Proximo pago" después de un espacio (si existen)
            if not df_proximos.empty:
                fila_proximos_start = fila_totales + 3 # Dejar 2 filas de espacio
                for i, (_, row) in enumerate(df_proximos.iterrows()):
                    current_row = fila_proximos_start + i
                    for col_idx, header in enumerate(headers, 1):
                        val = row.get(header)
                        if pd.notna(val):
                            ws.cell(row=current_row, column=col_idx).value = val
                    # Asegurar que diga "Proximo pago" al final si corresponde
                    if idx_observaciones:
                        ws.cell(row=current_row, column=idx_observaciones).value = "Proximo pago"

            # 6. Definir y aplicar estilos de cabecera
            color_fondo = PatternFill(start_color='69E2FF', end_color='69E2FF', fill_type='solid')
            fuente_cabecera = Font(bold=True)
            alineacion_centrada = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos a la fila de cabecera
            for cell in ws[1]:
                cell.fill = color_fondo
                cell.font = fuente_cabecera
                cell.alignment = alineacion_centrada
            
            # Configurar anchos de columna
            anchos_config = {
                "Observaciones": 44,
                "Date": 15,
                "Fecha del envío": 18,
                "Prorrateo Disputa": 20,
                "Prorrateo Normal": 20,
                "Neto despues de prorrateo": 20,
                "Fecha_pago": 15,
                "Order Id Paypal": 20,
                "Valoración flete": 20,
                "Diferencia": 15,
                "Número guía": 20,
                "Invoice Numbers": 18
            }
            
            for col_idx, col_name in enumerate(headers, 1):
                if col_name in anchos_config:
                    letra_col = get_column_letter(col_idx)
                    ws.column_dimensions[letra_col].width = anchos_config[col_name]
            
            wb.save(archivo)
            self.logger.info("Archivo guardado con totales, fórmulas dinámicas y registros parciales.")
        
        except Exception as e:
            self.logger.error(f"ERROR AL GUARDAR EL ARCHIVO EXCEL FINAL CON DISEÑO: {str(e)}")
            raise

# ============================================================================
# FASE 4: GESTIÓN DE PDFs
# ============================================================================

class GestorPDFs:
    """Maneja búsqueda, extracción y validación de PDFs"""
    
    def __init__(self, rutas_busqueda: List[Path]):
        self.rutas_busqueda = rutas_busqueda
        self.logger = logging.getLogger(__name__)
    
    def buscar_documentos_por_patron(self, dato_columna: str, prefijo: str = "") -> List[Path]:
        """
        Busca PDFs que coincidan con un dato y un prefijo opcional (ej: 'Guia ')
        """
        pdfs_encontrados = []
        try:
            # Limpiar y separar si hay varios datos (comas o puntos y coma)
            datos = [d.strip() for d in str(dato_columna).replace(';', ',').split(',') if d.strip()]
            
            if not datos:
                return []
            
            # Crear términos de búsqueda: normal y sin espacios para mayor flexibilidad
            terminos_busqueda = []
            for d in datos:
                base = d.lower()
                # Término tal cual
                terminos_busqueda.append(f"{prefijo.lower()}{base}".strip())
                # Término sin espacios internos (ej: '21 7694 0905' -> '2176940905')
                if " " in base:
                    base_sin_espacios = base.replace(" ", "")
                    terminos_busqueda.append(f"{prefijo.lower()}{base_sin_espacios}".strip())

            self.logger.info(f"Buscando documentos con patrones: {', '.join(terminos_busqueda)}...")
            
            for termino in terminos_busqueda:
                for ruta in self.rutas_busqueda:
                    if not ruta.exists():
                        continue
                    
                    try:
                        for pdf_file in ruta.rglob("*.pdf"):
                            nombre_archivo = pdf_file.name.lower()
                            # Búsqueda flexible: el término está en el nombre o el nombre sin espacios contiene el término
                            if termino in nombre_archivo or termino in nombre_archivo.replace(" ", ""):
                                self.logger.info(f"¡DOCUMENTO ENCONTRADO!: {pdf_file.name}")
                                pdfs_encontrados.append(pdf_file)
                    except Exception as e:
                        self.logger.error(f"Error al buscar '{termino}' en {ruta}: {e}")
            
            return list(set(pdfs_encontrados))
        
        except Exception as e:
            self.logger.error(f"ERROR DURANTE LA BÚSQUEDA DE DOCUMENTOS PARA {dato_columna}: {str(e)}")
            return pdfs_encontrados

    def extraer_fecha_pdf(self, pdf_path: Path) -> Optional[datetime]:
        """
        Extrae la fecha de un PDF (prioriza formato YYYY/MM/DD)
        """
        try:
            if not pdf_path.exists():
                return None

            doc = fitz.open(pdf_path)
            texto_completo = ""
            
            for page_num in range(min(3, len(doc))):
                page = doc[page_num]
                texto_completo += page.get_text()
            
            doc.close()
            
            if not texto_completo.strip():
                self.logger.warning(f"El PDF {pdf_path.name} no tiene texto extraíble.")
                return None

            import re
            # Patrones de fecha, priorizando YYYY/MM/DD o YYYY-MM-DD
            patrones = [
                r'\b(\d{4})[/-](\d{1,2})[/-](\d{1,2})\b',  # YYYY/MM/DD o YYYY-MM-DD
                r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b',  # DD/MM/YYYY o DD-MM-YYYY
                r'\b(\d{1,2})\s+(?:de\s+)?(\w+)\s+(?:de\s+)?(\d{4})\b'  # DD de Mes de YYYY
            ]
            
            meses_esp = {
                'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
                'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
                'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
            }
            
            for patron in patrones:
                matches = re.findall(patron, texto_completo, re.IGNORECASE)
                
                if matches:
                    match = matches[0]
                    
                    try:
                        # Intentar parsear según el patrón
                        if len(match) == 3:
                            if match[2].isdigit() and len(match[2]) == 4:
                                # DD/MM/YYYY
                                dia = int(match[0])
                                mes = int(match[1]) if match[1].isdigit() else meses_esp.get(match[1].lower(), 0)
                                año = int(match[2])
                            else:
                                # YYYY-MM-DD
                                año = int(match[0])
                                mes = int(match[1])
                                dia = int(match[2])
                            
                            if mes == 0: continue # Mes no reconocido

                            fecha = datetime(año, mes, dia)
                            self.logger.info(f"Fecha extraída de {pdf_path.name}: {fecha.strftime('%Y-%m-%d')}")
                            return fecha
                    
                    except (ValueError, KeyError):
                        continue
            
            self.logger.warning(f"No se detectó ninguna fecha válida en el texto de: {pdf_path.name}")
            return None
        
        except Exception as e:
            self.logger.error(f"ERROR AL EXTRAER FECHA DEL PDF {pdf_path.name}: {str(e)}")
            return None
    
    def procesar_documentos_soporte(self, df: pd.DataFrame, carpeta_soporte: Path) -> pd.DataFrame:
        """
        Busca, mueve y valida PDFs siguiendo el flujo: Copiar -> Clasificar -> Validar
        """
        try:
            col_invoices = 'Invoice Numbers'
            col_guias = 'Número guía'
            
            if col_invoices not in df.columns:
                self.logger.error(f"ERROR: Columna '{col_invoices}' no encontrada.")
                return df
            
            if 'Observaciones' not in df.columns:
                df['Observaciones'] = ""
            
            total_procesar = len(df)
            self.logger.info(f"Iniciando flujo de soportes para {total_procesar} registros...")
            
            for idx, row in df.iterrows():
                # Si el registro ya está marcado como "Proximo pago", lo saltamos
                if str(row.get('Observaciones', '')).strip() == "Proximo pago":
                    self.logger.info(f"Saltando registro {idx} (marcado como Proximo pago)")
                    continue

                invoice_val = str(row[col_invoices]).strip()
                guia_val = str(row.get(col_guias, "")).strip()
                
                # --- FASE 1: BUSCAR Y COPIAR TODO ---
                documentos_encontrados = []
                
                # A. Buscar por Invoice (Factura)
                if invoice_val and invoice_val.lower() != 'nan':
                    documentos_encontrados.extend(self.buscar_documentos_por_patron(invoice_val))
                    # B. Buscar Guía por el valor de Invoice (ej: "Guia COUR3515")
                    documentos_encontrados.extend(self.buscar_documentos_por_patron(invoice_val, prefijo="Guia "))
                
                # C. Buscar por Número de guía (ej: "Guia 21 7696...")
                if guia_val and guia_val.lower() != 'nan':
                    documentos_encontrados.extend(self.buscar_documentos_por_patron(guia_val, prefijo="Guia "))
                    # También buscar el número de guía solo por si acaso
                    documentos_encontrados.extend(self.buscar_documentos_por_patron(guia_val))
                
                # Eliminar duplicados de búsqueda
                documentos_encontrados = list(set(documentos_encontrados))
                
                # Copiar a carpeta soporte
                for doc_path in documentos_encontrados:
                    destino = carpeta_soporte / doc_path.name
                    if not destino.exists():
                        shutil.copy2(doc_path, destino)
                        self.logger.info(f"Copiado a Soporte: {doc_path.name}")

                # --- FASE 2: CLASIFICAR LO GUARDADO ---
                archivos_en_soporte = list(carpeta_soporte.glob("*.pdf"))
                
                tiene_factura = False
                tiene_guia = False
                guia_encontrada_path = None
                
                # Normalizamos valores para comparación
                inv_clean = invoice_val.lower().replace(" ", "")
                gui_clean = guia_val.lower().replace(" ", "")

                for archivo in archivos_en_soporte:
                    nombre_low = archivo.name.lower()
                    nombre_clean = nombre_low.replace(" ", "")
                    
                    # 1. Es una GUÍA si:
                    # - El nombre contiene 'guia' Y el valor de invoice (ej: Guia COUR3515...)
                    # - O el nombre contiene el número de guía (ej: Guia 21 7696... o 217696...)
                    es_guia_por_invoice = (inv_clean and inv_clean != 'nan' and "guia" in nombre_clean and inv_clean in nombre_clean)
                    es_guia_por_numero = (gui_clean and gui_clean != 'nan' and gui_clean in nombre_clean)
                    
                    if es_guia_por_invoice or es_guia_por_numero:
                        tiene_guia = True
                        guia_encontrada_path = archivo
                        continue # Si es guía, no la evaluamos como factura

                    # 2. Es una FACTURA si:
                    # - Contiene el invoice_val Y NO contiene la palabra 'guia'
                    if inv_clean and inv_clean != 'nan' and inv_clean in nombre_clean:
                        if "guia" not in nombre_low:
                            tiene_factura = True

                # --- FASE 3: VALIDAR Y ASIGNAR OBSERVACIONES ---
                fecha_coincide = True
                fecha_anterior_str = ""
                
                # Validar fecha si hay guía
                if tiene_guia and guia_encontrada_path:
                    fecha_pdf = self.extraer_fecha_pdf(guia_encontrada_path)
                    if fecha_pdf:
                        # Guardamos la fecha que estaba originalmente en el Excel (siempre está lleno)
                        fecha_ref = row.get('Fecha del envío')
                        
                        if pd.notna(fecha_ref):
                            try:
                                dt_ref = pd.to_datetime(fecha_ref)
                                fecha_anterior_str = dt_ref.strftime('%y/%m/%d')
                                # Comparar mes y año entre la guía (PDF) y el Excel original
                                if fecha_pdf.month != dt_ref.month or fecha_pdf.year != dt_ref.year:
                                    fecha_coincide = False
                            except:
                                fecha_anterior_str = str(fecha_ref)
                                fecha_coincide = False

                        # SIEMPRE actualizamos con la fecha del PDF si la encontramos
                        df.at[idx, 'Fecha del envío'] = fecha_pdf.date()
                    else:
                        self.logger.warning(f"No se pudo extraer fecha del PDF: {guia_encontrada_path.name}")

                # Determinar observación final basada en los casos solicitados
                observacion_final = ""
                msg_fecha_anterior = f"Fecha anterior registrada {fecha_anterior_str}"
                
                if not tiene_factura and not tiene_guia:
                    observacion_final = "Faltan ambos documentos"
                elif tiene_factura and not tiene_guia:
                    observacion_final = "Falta la guia de transporte"
                elif not tiene_factura and tiene_guia:
                    if not fecha_coincide:
                        observacion_final = f"Falta la factura comercial y {msg_fecha_anterior}"
                    else:
                        observacion_final = "Falta la factura comercial"
                elif tiene_factura and tiene_guia:
                    if not fecha_coincide:
                        observacion_final = msg_fecha_anterior
                    else:
                        observacion_final = "Soportes OK"

                df.at[idx, 'Observaciones'] = observacion_final

            self.logger.info("Procesamiento de soportes finalizado con el nuevo flujo.")
            return df
        
        except Exception as e:
            self.logger.error(f"ERROR CRÍTICO EN PROCESAMIENTO DE DOCUMENTOS: {str(e)}")
            return df

# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """Función principal que orquesta todo el proceso"""
    
    logger = configurar_logging()
    logger.info("=" * 80)
    logger.info("INICIANDO SISTEMA DE AUTOMATIZACIÓN DE PAGOS PAYPAL")
    logger.info("=" * 80)
    
    fase_actual = "Inicialización"
    try:
        # ========================================
        # PASO 1: Gestión de Carpetas
        # ========================================
        fase_actual = "Gestión de Carpetas"
        logger.info("\n[PASO 1] Identificando pago pendiente o nueva carpeta...")
        
        gestor_carpetas = GestorCarpetas(Config.BASE_PAYPAL)
        numero_pago = gestor_carpetas.obtener_pago_pendiente_o_siguiente()
        carpeta_pago, carpeta_soporte = gestor_carpetas.crear_estructura_pago(numero_pago)
        
        logger.info(f"Número de pago a procesar: {numero_pago}")
        logger.info(f"Carpeta destino: {carpeta_pago}")
        
        # ========================================
        # PASO 2: Obtener Reporte (SAP con apoyo manual)
        # ========================================
        fase_actual = "Obtención de Reporte"
        logger.info("\n[PASO 2] Obteniendo reporte...")
        
        descargador = DescargadorSAP()
        archivo_descarga = descargador.descargar_reporte_sap(numero_pago)
        
        if not archivo_descarga:
            logger.warning(f"No se pudo descargar automáticamente. Buscando archivo 'pago {numero_pago}' en Descargas...")
            procesador = ProcesadorExcel()
            archivo_descarga = procesador.buscar_archivo_pago_en_descargas(numero_pago)
        
        if not archivo_descarga:
            logger.error(f"FALLO: No se encontró el reporte para el pago {numero_pago}")
            logger.info("Asegúrese de que el archivo esté en Descargas con el nombre correcto.")
            return 1
        
        # ========================================
        # PASO 3: Procesar Excel - Primera Hoja
        # ========================================
        fase_actual = "Procesamiento de Excel (Hoja 1)"
        logger.info("\n[PASO 3] Procesando archivo Excel...")
        
        procesador = ProcesadorExcel() # Asegurar que esté inicializado
        
        # Mover archivo a carpeta del pago
        archivo_principal = procesador.mover_y_renombrar_descarga(
            archivo_descarga, carpeta_pago, numero_pago
        )
        
        # Reorganizar columnas (mover Referencia al inicio)
        procesador.reorganizar_columnas_primera_hoja(archivo_principal)
        
        # ========================================
        # PASO 4: Crear Segunda Hoja
        # ========================================
        fase_actual = "Procesamiento de Excel (Hoja 2)"
        logger.info("\n[PASO 4] Creando segunda hoja desde maestro...")
        
        if not Config.RUTA_MAESTRO.exists():
            logger.error(f"ARCHIVO MAESTRO NO ENCONTRADO EN: {Config.RUTA_MAESTRO}")
            return 1
        
        df_segunda_hoja = procesador.crear_segunda_hoja(
            archivo_principal, Config.RUTA_MAESTRO
        )
        
        # Calcular Comparación flete y Resultado comparación
        df_segunda_hoja = procesador.calcular_mon_grupo_y_diferencia(
            archivo_principal, df_segunda_hoja
        )
        
        # ========================================
        # PASO 5: Procesar PDFs
        # ========================================
        fase_actual = "Procesamiento de Soporte PDF"
        logger.info("\n[PASO 5] Buscando y validando documentos PDF...")
        
        gestor_pdfs = GestorPDFs(Config.RUTAS_PDF)
        df_segunda_hoja = gestor_pdfs.procesar_documentos_soporte(
            df_segunda_hoja, carpeta_soporte
        )
        
        # ========================================
        # PASO 6: Guardar Excel Final
        # ========================================
        fase_actual = "Guardado Final"
        logger.info("\n[PASO 6] Guardando archivo Excel final...")
        
        procesador.guardar_excel_con_dos_hojas(archivo_principal, df_segunda_hoja)
        
        # ========================================
        # RESUMEN FINAL
        # ========================================
        logger.info("\n" + "=" * 80)
        logger.info("PROCESO COMPLETADO EXITOSAMENTE")
        logger.info("=" * 80)
        logger.info(f"Pago #: {numero_pago}")
        logger.info(f"Ubicación: {carpeta_pago}")
        logger.info(f"Archivo Excel: {archivo_principal.name}")
        logger.info(f"Documentos en Soporte: {len(list(carpeta_soporte.glob('*.pdf')))}")
        logger.info("=" * 80)
        
        # Mostrar advertencias si hay
        registros_sin_docs = df_segunda_hoja[
            df_segunda_hoja['Observaciones'].str.contains('Falta|Sin|DOCUMENTO NO ENCONTRADO', na=False)
        ]
        
        if len(registros_sin_docs) > 0:
            logger.warning(f"\nADVERTENCIA: {len(registros_sin_docs)} registros con observaciones o sin documentos")
            logger.warning("Por favor, revise la columna 'Observaciones' en la segunda hoja del Excel generado.")
        
        return 0
    
    except KeyboardInterrupt:
        logger.warning("\nProceso interrumpido por el usuario (Ctrl+C)")
        return 1
    
    except Exception as e:
        logger.error("\n" + "!" * 80)
        logger.error(f"ERROR CRÍTICO DURANTE LA FASE: {fase_actual.upper()}")
        logger.error(f"DETALLE DEL ERROR: {str(e)}")
        logger.error("!" * 80)
        logger.debug("Traceback completo:", exc_info=True)
        return 1

if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as e:
        print(f"\nError fatal: {e}")
        input("\nPresione Enter para salir...")
        sys.exit(1)
