"""
VERIFICADOR Y ACTUALIZADOR DE SOPORTES - VERSIÓN MEJORADA
Busca documentos en OneDrive, los copia a Soporte y actualiza Excel automáticamente
"""

import logging
import shutil
from pathlib import Path
from typing import Optional, Tuple, List, Dict
import pandas as pd
from dataclasses import dataclass, field
from enum import Enum


class EstadoSoporte(Enum):
    """Estados posibles de un soporte"""
    COMPLETO = " Soportes OK"
    FALTA_FACTURA = "Falta la factura comercial"
    FALTA_GUIA = "Falta la guia de transporte"
    FALTA_AMBOS = "Faltan ambos documentos"
    FECHA_ANTERIOR = " Fecha anterior registrada"


@dataclass
class ResultadoVerificacion:
    """Resultado de la verificación y actualización de un pago"""
    numero_pago: int
    registros_totales: int
    registros_con_observaciones: int
    documentos_copiados: int
    observaciones_actualizadas: int
    estado_general: str
    detalles: List[Dict] = field(default_factory=list)
    archivos_copiados: List[Dict] = field(default_factory=list)
    cambios_realizados: List[Dict] = field(default_factory=list)
    archivo_excel: Optional[Path] = None
    carpeta_soporte: Optional[Path] = None


class VerificadorActualizadorSoportes:
    """
    Verifica soportes de pagos, copia documentos de OneDrive a Soporte
    y actualiza automáticamente las observaciones en el Excel
    """
    
    def __init__(self, rutas_pdf: List[Path]):
        """
        Inicializa el verificador/actualizador
        
        Args:
            rutas_pdf: Lista de rutas donde buscar PDFs (OneDrive, etc.)
        """
        self.rutas_pdf = rutas_pdf
        self.logger = logging.getLogger(__name__)
    
    def obtener_pagos_existentes(self, base_paypal: Path) -> List[Tuple[int, Path]]:
        """Obtiene lista de carpetas de pagos existentes"""
        pagos = []
        try:
            if not base_paypal.exists():
                self.logger.error(f"Ruta base no existe: {base_paypal}")
                return pagos
            
            for item in base_paypal.iterdir():
                if item.is_dir() and item.name.startswith("Pago #"):
                    try:
                        numero = int(item.name.replace("Pago #", ""))
                        pagos.append((numero, item))
                    except ValueError:
                        continue
            
            pagos.sort()
            self.logger.info(f"Se encontraron {len(pagos)} pagos: {[p[0] for p in pagos]}")
            return pagos
        
        except Exception as e:
            self.logger.error(f"Error al obtener pagos: {e}")
            return pagos
    
    def obtener_archivo_excel_pago(self, carpeta_pago: Path) -> List[Path]:
        """Busca y retorna una lista de candidatos a archivo Excel ordenados por relevancia"""
        try:
            extensiones = ["*.xlsx", "*.xlsm", "*.xls"]
            candidatos = []
            
            for ext in extensiones:
                for archivo in carpeta_pago.glob(ext):
                    if archivo.name.startswith("~$") or "Soporte" in archivo.name:
                        continue
                    candidatos.append(archivo)
            
            if not candidatos:
                return []

            # CRITERIOS DE ORDENACIÓN (Relevancia):
            # 1. Archivos que empiezan por "EXPORT" o contienen "Pago"
            # 2. Tamaño del archivo (los reportes reales pesan > 1MB)
            # 3. Fecha de modificación
            
            def prioridad(f: Path):
                score = 0
                name_lower = f.name.lower()
                size_mb = f.stat().st_size / (1024 * 1024)
                
                if name_lower.startswith("export"): score += 100
                if "pago" in name_lower: score += 50
                if size_mb > 0.5: score += 30  # Más de 500KB es buena señal
                
                # Sumar un pequeño factor por fecha para desempatar
                score += (f.stat().st_mtime / 1_000_000_000) 
                return score

            candidatos.sort(key=prioridad, reverse=True)
            self.logger.info(f"Candidatos encontrados: {[f.name for f in candidatos]}")
            return candidatos
            
        except Exception as e:
            self.logger.error(f"Error buscando candidatos Excel en {carpeta_pago}: {e}")
            return []
    
    def leer_segunda_hoja_excel(self, archivo_excel: Path) -> Optional[pd.DataFrame]:
        """Lee la segunda hoja del Excel (Reporte Procesado) con dtype object para evitar truncamiento de IDs"""
        try:
            # Intentar obtener los nombres de las hojas primero
            xl = pd.ExcelFile(archivo_excel)
            hojas = xl.sheet_names
            
            # Nombre esperado
            nombre_esperado = 'Reporte Procesado'
            
            if nombre_esperado in hojas:
                df = pd.read_excel(archivo_excel, sheet_name=nombre_esperado, engine='openpyxl', dtype=object)
            elif len(hojas) >= 2:
                # Si no existe el nombre, pero hay al menos 2 hojas, intentar con la segunda hoja por índice
                self.logger.warning(f"No se encontró la hoja '{nombre_esperado}'. Intentando con la segunda hoja: '{hojas[1]}'")
                df = pd.read_excel(archivo_excel, sheet_name=1, engine='openpyxl', dtype=object)
            else:
                self.logger.error(f"El archivo Excel {archivo_excel.name} solo tiene una hoja: {hojas}. Se requieren al menos 2.")
                return None

            # CORRECCIÓN DE FECHAS: Asegurar formato dd/mm/aaaa y quitar horas
            columnas_fecha = ["Date", "Fecha del envío", "Fecha_pago"]
            for col in columnas_fecha:
                if col in df.columns:
                    # Convertir a datetime y luego a string con formato dd/mm/yyyy
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y')
                    # Reemplazar 'NaT' por vacío
                    df[col] = df[col].replace('NaT', "")

            self.logger.info(f"Se leyeron {len(df)} registros de {archivo_excel.name} (fechas corregidas)")
            return df
        except Exception as e:
            self.logger.error(f"Error leyendo Excel {archivo_excel}: {e}")
            return None
    
    def buscar_documentos_por_patron(self, dato_columna: str, prefijo: str = "") -> List[Path]:
        """Busca PDFs que coincidan con un dato y un prefijo opcional"""
        pdfs_encontrados = []
        try:
            datos = [d.strip() for d in str(dato_columna).replace(';', ',').split(',') if d.strip()]
            
            if not datos:
                return []
            
            terminos_busqueda = []
            for d in datos:
                base = d.lower()
                terminos_busqueda.append(f"{prefijo.lower()}{base}".strip())
                if " " in base:
                    base_sin_espacios = base.replace(" ", "")
                    terminos_busqueda.append(f"{prefijo.lower()}{base_sin_espacios}".strip())

            self.logger.info(f"Buscando: {', '.join(terminos_busqueda)}...")
            
            for termino in terminos_busqueda:
                for ruta in self.rutas_pdf:
                    if not ruta.exists():
                        continue
                    
                    try:
                        for pdf_file in ruta.rglob("*.pdf"):
                            nombre_archivo = pdf_file.name.lower()
                            if termino in nombre_archivo or termino in nombre_archivo.replace(" ", ""):
                                self.logger.info(f"OK ENCONTRADO: {pdf_file.name}")
                                pdfs_encontrados.append(pdf_file)
                    except Exception as e:
                        self.logger.error(f"Error buscando '{termino}' en {ruta}: {e}")
            
            return list(set(pdfs_encontrados))
        
        except Exception as e:
            self.logger.error(f"ERROR EN BÚSQUEDA: {str(e)}")
            return pdfs_encontrados
    
    def copiar_documentos_a_soporte(self, 
                                   documentos: List[Path], 
                                   carpeta_soporte: Path) -> List[Dict]:
        """
        Copia documentos encontrados a la carpeta Soporte
        
        Returns:
            Lista de dicts con info de archivos copiados
        """
        archivos_copiados = []
        
        try:
            carpeta_soporte.mkdir(parents=True, exist_ok=True)
            
            for doc_path in documentos:
                try:
                    destino = carpeta_soporte / doc_path.name
                    
                    # Si ya existe, no copiar
                    if destino.exists():
                        self.logger.info(f"AVISO: Ya existe: {doc_path.name}")
                        continue
                    
                    # Copiar archivo
                    shutil.copy2(doc_path, destino)
                    self.logger.info(f"OK Copiado a Soporte: {doc_path.name}")
                    
                    archivos_copiados.append({
                        'nombre': doc_path.name,
                        'origen': str(doc_path),
                        'destino': str(destino),
                        'tamaño': doc_path.stat().st_size
                    })
                
                except Exception as e:
                    self.logger.error(f"Error copiando {doc_path.name}: {e}")
            
            return archivos_copiados
        
        except Exception as e:
            self.logger.error(f"Error en copiar_documentos: {e}")
            return archivos_copiados
    
    def obtener_documentos_en_soporte(self, carpeta_soporte: Path) -> Dict[str, List[Path]]:
        """Obtiene documentos organizados por tipo en la carpeta Soporte"""
        documentos = {'guias': [], 'facturas': []}
        
        try:
            if not carpeta_soporte.exists():
                return documentos
            
            for pdf in carpeta_soporte.glob("*.pdf"):
                nombre = pdf.name.lower()
                if "guia" in nombre:
                    documentos['guias'].append(pdf)
                else:
                    documentos['facturas'].append(pdf)
            
            return documentos
        
        except Exception as e:
            self.logger.error(f"Error leyendo documentos: {e}")
            return documentos
    
    def buscar_archivo_en_soporte(self, 
                                 documentos_soporte: Dict, 
                                 numero_referencia: str, 
                                 tipo: str = 'factura',
                                 referencia_alternativa: Optional[str] = None) -> bool:
        """
        Busca si un documento existe en los archivos de Soporte
        
        Args:
            documentos_soporte: Diccionario de documentos organizados por tipo
            numero_referencia: Número principal a buscar (Invoice o Guía)
            tipo: 'factura' o 'guia'
            referencia_alternativa: Opcional, segundo número para intentar si falla el primero
        """
        def _buscar(ref):
            if not ref or str(ref).lower() == 'nan' or str(ref).strip() == "":
                return False
            
            ref_clean = str(ref).lower().replace(" ", "")
            lista_documentos = documentos_soporte.get('guias' if tipo == 'guia' else 'facturas', [])
            
            for doc in lista_documentos:
                nombre_clean = doc.name.lower().replace(" ", "")
                # Búsqueda flexible: el número está en el nombre o viceversa
                if ref_clean in nombre_clean or (len(ref_clean) > 5 and nombre_clean in ref_clean):
                    return True
            return False

        # Intentar con la referencia principal
        if _buscar(numero_referencia):
            return True
            
        # Intentar con la referencia alternativa si existe
        if referencia_alternativa and _buscar(referencia_alternativa):
            return True
        
        return False
    
    def analizar_observaciones_registro(self, row: pd.Series) -> Dict:
        """Analiza qué documentos faltan según las observaciones"""
        observaciones = str(row.get('Observaciones', '')).strip()
        invoice = str(row.get('Invoice Numbers', '')).strip()
        guia = str(row.get('Número guía', '')).strip()
        
        resultado = {
            'falta_factura': False,
            'falta_guia': False,
            'observacion_original': observaciones,
            'invoice': invoice,
            'guia': guia,
            'corregir_excel': False # Flag para forzar guardado si se corrigieron fechas
        }
        
        if not observaciones or observaciones.lower() == 'soportes ok':
            return resultado
        
        obs_lower = observaciones.lower()
        
        if 'falta' in obs_lower:
            if 'factura' in obs_lower:
                resultado['falta_factura'] = True
            if 'guia' in obs_lower or 'guía' in obs_lower:
                resultado['falta_guia'] = True
            if 'ambos' in obs_lower:
                resultado['falta_factura'] = True
                resultado['falta_guia'] = True
        
        return resultado
    
    def determinar_nueva_observacion(self, 
                                     falta_factura: bool, 
                                     falta_guia: bool,
                                     factura_encontrada: bool,
                                     guia_encontrada: bool) -> str:
        """Determina la nueva observación basada en estado actual"""
        
        # Si no faltaba nada
        if not falta_factura and not falta_guia:
            return "Soportes OK"
        
        # Ambos documentos faltaban
        if falta_factura and falta_guia:
            if factura_encontrada and guia_encontrada:
                return "Soportes OK"
            elif factura_encontrada:
                return "Falta la guia de transporte"
            elif guia_encontrada:
                return "Falta la factura comercial"
            else:
                return "Faltan ambos documentos"
        
        # Solo faltaba factura
        if falta_factura:
            if factura_encontrada:
                return "Soportes OK"
            else:
                return "Falta la factura comercial"
        
        # Solo faltaba guía
        if falta_guia:
            if guia_encontrada:
                return "Soportes OK"
            else:
                return "Falta la guia de transporte"
        
        return "Soportes OK"
    
    def actualizar_excel_con_nuevas_observaciones(self, 
                                                  archivo_excel: Path,
                                                  df_actualizado: pd.DataFrame) -> bool:
        """Guarda el DataFrame actualizado en el Excel"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            
            self.logger.info(f"Guardando cambios en {archivo_excel.name}...")
            
            # Guardar con pandas
            with pd.ExcelWriter(archivo_excel, engine='openpyxl', mode='a', 
                               if_sheet_exists='replace') as writer:
                df_actualizado.to_excel(writer, sheet_name='Reporte Procesado', index=False)
            
            # Aplicar estilos con openpyxl
            wb = load_workbook(archivo_excel)
            ws = wb['Reporte Procesado']
            
            # Estilo cabecera
            color_fondo = PatternFill(start_color='69E2FF', end_color='69E2FF', fill_type='solid')
            fuente_cabecera = Font(bold=True)
            alineacion_centrada = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.fill = color_fondo
                cell.font = fuente_cabecera
                cell.alignment = alineacion_centrada
            
            # Anchos de columna
            anchos_config = {
                "Observaciones": 44,
                "Date": 15,
                "Fecha del envío": 18,
                "Invoice Numbers": 18,
                "Número guía": 20
            }
            
            headers = [cell.value for cell in ws[1]]
            for col_idx, col_name in enumerate(headers, 1):
                if col_name in anchos_config:
                    letra_col = get_column_letter(col_idx)
                    ws.column_dimensions[letra_col].width = anchos_config[col_name]
            
            wb.save(archivo_excel)
            self.logger.info("Excel guardado con éxito")
            return True
        
        except Exception as e:
            self.logger.error(f"Error guardando Excel: {e}")
            return False
    
    def procesar_pago_completo(self, numero_pago: int, base_paypal: Path) -> ResultadoVerificacion:
        """
        Proceso COMPLETO:
        1. Busca documentos faltantes en OneDrive
        2. Los copia a carpeta Soporte
        3. Actualiza observaciones en Excel (Y corrige formatos de fecha)
        4. Retorna reporte de cambios
        """
        self.logger.info(f"\n{'='*70}")
        self.logger.info(f"PROCESANDO PAGO #{numero_pago}")
        self.logger.info(f"{'='*70}")
        
        carpeta_pago = base_paypal / f"Pago #{numero_pago}"
        
        if not carpeta_pago.exists():
            self.logger.error(f"Carpeta no existe: {carpeta_pago}")
            return ResultadoVerificacion(
                numero_pago=numero_pago,
                registros_totales=0,
                registros_con_observaciones=0,
                documentos_copiados=0,
                observaciones_actualizadas=0,
                estado_general="Carpeta no encontrada",
                detalles=[]
            )
        
        # Obtener lista de archivos Excel candidatos
        candidatos_excel = self.obtener_archivo_excel_pago(carpeta_pago)
        if not candidatos_excel:
            self.logger.error(f"No se encontró ningún archivo Excel en {carpeta_pago}")
            return ResultadoVerificacion(
                numero_pago=numero_pago,
                registros_totales=0,
                registros_con_observaciones=0,
                documentos_copiados=0,
                observaciones_actualizadas=0,
                estado_general="No se encontró archivo Excel",
                detalles=[]
            )
        
        # Intentar leer cada candidato hasta encontrar uno válido
        df = None
        archivo_excel = None
        
        for candidato in candidatos_excel:
            self.logger.info(f"Intentando leer candidato: {candidato.name}")
            df = self.leer_segunda_hoja_excel(candidato)
            if df is not None and not df.empty:
                archivo_excel = candidato
                self.logger.info(f"¡Candidato aceptado!: {archivo_excel.name}")
                break
            else:
                self.logger.warning(f"Candidato descartado (sin datos o sin hoja correcta): {candidato.name}")
        
        if df is None or df.empty:
            self.logger.error(f"Ninguno de los {len(candidatos_excel)} archivos Excel contiene datos válidos.")
            return ResultadoVerificacion(
                numero_pago=numero_pago,
                registros_totales=0,
                registros_con_observaciones=0,
                documentos_copiados=0,
                observaciones_actualizadas=0,
                estado_general="Excel sin datos válidos",
                detalles=[]
            )
        
        # Crear copia para actualizar
        df_actualizado = df.copy()
        
        # Carpeta soporte
        carpeta_soporte = carpeta_pago / "Soporte"
        carpeta_soporte.mkdir(parents=True, exist_ok=True)
        
        # PASO 1: BUSCAR Y COPIAR DOCUMENTOS
        self.logger.info("\nPASO 1: Buscando y copiando documentos...")
        todos_documentos_encontrados = []
        archivos_copiados = []
        
        for idx, row in df.iterrows():
            observacion = str(row.get('Observaciones', '')).strip()
            
            # Saltar si ya está completo PERO procesar igual para copiar PDFs si faltan físicamente
            invoice = str(row.get('Invoice Numbers', '')).strip()
            guia = str(row.get('Número guía', '')).strip()
            
            invoice = str(row.get('Invoice Numbers', '')).strip()
            guia = str(row.get('Número guía', '')).strip()
            
            # Buscar documentos
            documentos = []
            if invoice and invoice.lower() != 'nan':
                documentos.extend(self.buscar_documentos_por_patron(invoice))
                documentos.extend(self.buscar_documentos_por_patron(invoice, prefijo="Guia "))
            
            if guia and guia.lower() != 'nan':
                documentos.extend(self.buscar_documentos_por_patron(guia, prefijo="Guia "))
                documentos.extend(self.buscar_documentos_por_patron(guia))
            
            documentos = list(set(documentos))
            todos_documentos_encontrados.extend(documentos)
            
            # Copiar a Soporte
            copiados = self.copiar_documentos_a_soporte(documentos, carpeta_soporte)
            archivos_copiados.extend(copiados)
        
        # PASO 2: ANALIZAR Y ACTUALIZAR OBSERVACIONES
        self.logger.info("\nPASO 2: Analizando y actualizando observaciones...")
        
        documentos_en_soporte = self.obtener_documentos_en_soporte(carpeta_soporte)
        detalles = []
        cambios_realizados = []
        observaciones_actualizadas = 0
        registros_con_observaciones = 0
        
        for idx, row in df.iterrows():
            info_obs = self.analizar_observaciones_registro(row)
            observacion_original = info_obs['observacion_original']
            
            # Saltar registros sin observaciones
            if not observacion_original or observacion_original.lower() == 'soportes ok':
                continue
            
            registros_con_observaciones += 1
            
            # Buscar documentos en Soporte
            factura_encontrada = False
            guia_encontrada = False
            
            if info_obs['falta_factura']:
                factura_encontrada = self.buscar_archivo_en_soporte(
                    documentos_en_soporte, 
                    info_obs['invoice'], 
                    'factura',
                    referencia_alternativa=info_obs['guia']
                )
            
            if info_obs['falta_guia']:
                guia_encontrada = self.buscar_archivo_en_soporte(
                    documentos_en_soporte, 
                    info_obs['guia'], 
                    'guia',
                    referencia_alternativa=info_obs['invoice']
                )
            
            # Determinar nueva observación
            nueva_observacion = self.determinar_nueva_observacion(
                info_obs['falta_factura'],
                info_obs['falta_guia'],
                factura_encontrada,
                guia_encontrada
            )
            
            # Si cambió, actualizar
            cambio = False
            if nueva_observacion != observacion_original:
                df_actualizado.at[idx, 'Observaciones'] = nueva_observacion
                cambio = True
                observaciones_actualizadas += 1
                
                self.logger.info(
                    f"  Fila {idx+2}: {observacion_original} → {nueva_observacion}"
                )
                
                cambios_realizados.append({
                    'fila': idx + 2,
                    'observacion_anterior': observacion_original,
                    'observacion_nueva': nueva_observacion,
                    'invoice': info_obs['invoice'],
                    'guia': info_obs['guia']
                })
            
            detalle = {
                'fila': idx + 2,
                'invoice': info_obs['invoice'],
                'guia': info_obs['guia'],
                'observacion_original': observacion_original,
                'observacion_nueva': nueva_observacion,
                'cambio': cambio,
                'falta_factura': info_obs['falta_factura'],
                'falta_guia': info_obs['falta_guia'],
                'factura_encontrada': factura_encontrada,
                'guia_encontrada': guia_encontrada
            }
            detalles.append(detalle)
        
        # PASO 3: GUARDAR EXCEL ACTUALIZADO
        self.logger.info("\nPASO 3: Guardando Excel actualizado y corrigiendo formatos de fecha...")
        
        # Siempre guardamos el Excel para asegurar que las correcciones de fecha se apliquen
        exito = self.actualizar_excel_con_nuevas_observaciones(archivo_excel, df_actualizado)
        if not exito:
            self.logger.error("Error al guardar Excel")
        
        # Estado general
        if len(archivos_copiados) > 0:
            estado_general = (
                f"OK {len(archivos_copiados)} archivos copiados, "
                f"fechas corregidas y {observaciones_actualizadas} observaciones"
            )
        else:
            estado_general = f"OK Fechas corregidas y {observaciones_actualizadas} observaciones actualizadas"
        
        self.logger.info(f"\n{'='*70}")
        self.logger.info(f"RESUMEN - Pago #{numero_pago}")
        self.logger.info(f"{'='*70}")
        self.logger.info(f"Registros totales: {len(df)}")
        self.logger.info(f"Con observaciones: {registros_con_observaciones}")
        self.logger.info(f"Archivos copiados: {len(archivos_copiados)}")
        self.logger.info(f"Observaciones actualizadas: {observaciones_actualizadas}")
        self.logger.info(f"Estado: {estado_general}")
        
        return ResultadoVerificacion(
            numero_pago=numero_pago,
            registros_totales=len(df),
            registros_con_observaciones=registros_con_observaciones,
            documentos_copiados=len(archivos_copiados),
            observaciones_actualizadas=observaciones_actualizadas,
            estado_general=estado_general,
            detalles=detalles,
            archivos_copiados=archivos_copiados,
            cambios_realizados=cambios_realizados,
            archivo_excel=archivo_excel,
            carpeta_soporte=carpeta_soporte
        )
    
    def generar_reporte_completo(self, resultado: ResultadoVerificacion) -> str:
        """Genera un reporte detallado de todo lo realizado"""
        reporte = "\n" + "="*80 + "\n"
        reporte += f"REPORTE COMPLETO - PAGO #{resultado.numero_pago}\n"
        reporte += "="*80 + "\n\n"
        
        reporte += f"ESTADO GENERAL: {resultado.estado_general}\n\n"
        
        reporte += f"ESTADÍSTICAS:\n"
        reporte += f"  Registros totales: {resultado.registros_totales}\n"
        reporte += f"  Con observaciones: {resultado.registros_con_observaciones}\n"
        reporte += f"  Archivos copiados: {resultado.documentos_copiados}\n"
        reporte += f"  Observaciones actualizadas: {resultado.observaciones_actualizadas}\n\n"
        
        if resultado.archivos_copiados:
            reporte += f"ARCHIVOS COPIADOS A SOPORTE ({len(resultado.archivos_copiados)}):\n"
            reporte += "-" * 80 + "\n"
            for archivo in resultado.archivos_copiados:
                reporte += f"  • {archivo['nombre']}\n"
                reporte += f"    De: {archivo['origen']}\n"
                reporte += f"    A: {archivo['destino']}\n"
            reporte += "\n"
        
        if resultado.cambios_realizados:
            reporte += f"OBSERVACIONES ACTUALIZADAS ({len(resultado.cambios_realizados)}):\n"
            reporte += "-" * 80 + "\n"
            for cambio in resultado.cambios_realizados:
                reporte += f"\n  Fila {cambio['fila']}:\n"
                reporte += f"    Invoice: {cambio['invoice']}\n"
                reporte += f"    Guía: {cambio['guia']}\n"
                reporte += f"    ANTES: {cambio['observacion_anterior']}\n"
                reporte += f"    DESPUÉS: {cambio['observacion_nueva']}\n"
            reporte += "\n"
        
        if resultado.detalles:
            reporte += f"DETALLE POR REGISTRO:\n"
            reporte += "-" * 80 + "\n"
            for detalle in resultado.detalles:
                reporte += f"\n  Fila {detalle['fila']}:\n"
                reporte += f"    Invoice: {detalle['invoice']}\n"
                reporte += f"    Guía: {detalle['guia']}\n"
                reporte += f"    Observación: {detalle['observacion_original']}\n"
                if detalle['falta_factura']:
                    reporte += f"    Factura: {'✅ ENCONTRADA' if detalle['factura_encontrada'] else 'FALTA'}\n"
                if detalle['falta_guia']:
                    reporte += f"    Guía: {'✅ ENCONTRADA' if detalle['guia_encontrada'] else 'FALTA'}\n"
                reporte += f"    Nueva observación: {detalle['observacion_nueva']}\n"
        
        reporte += "\n" + "="*80 + "\n"
        return reporte